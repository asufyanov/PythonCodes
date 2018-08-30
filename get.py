import urllib.request
import json
from openpyxl import load_workbook
from openpyxl import Workbook
#url = 'https://jsonplaceholder.typicode.com/users'

urlStart = 'https://api-metrika.yandex.ru/stat/v1/data?'

metrics = "metrics=ym:s:goal23489735reaches"
dimensions = "&dimensions=ym:s:goal"
date1 = "&date1=2018-08-20"
date2 = "&date2=2018-08-26"
ids = "&ids=10575199"
token = "&oauth_token=AQAAAAAfG7AEAAUtKjM8zFheiEHQpFBiD0qG6Kg"
accuracy = "&accuracy=high"



#req = urllib.request.Request(url)

d1 = input("Введите начальную дату YYYY-MM-DD")
d2 = input("Введите конечную дату YYYY-MM-DD")

date1 = "&date1="+d1
date2 = "&date2="+d2

colNames = []
colGoalReached = []
colUsers = []


goal = []
goal.append(29216499)
goal.append(29216504)
goal.append(29216964)
goal.append(29216969)
goal.append(29216974)
goal.append(40146673)
goal.append(40146715)
goal.append(40146754)
goal.append(40146784)
goal.append(40295419)
goal.append(40295425)
goal.append(40295458)

goalStart = "ym:s:goal"
goalEnd = "reaches"

for i in goal:
    metrics = "metrics=ym:s:goal"+str(i)+"reaches,ym:s:users"
    url = urlStart + metrics + dimensions + date1 + date2 + ids + token + accuracy
    print (url)
    req = urllib.request.Request(url)
    q = urllib.request.urlopen(req).read()
    cont = json.loads(q.decode('utf-8'))
    print (cont['data'][0]['dimensions'][0]['name'])
    colNames.append(cont['data'][0]['dimensions'][0]['name'])
    print (cont['data'][0]['metrics'][0], cont['data'][0]['metrics'][1])
    colGoalReached.append(cont['data'][0]['metrics'][0])
    colUsers.append(cont['data'][0]['metrics'][1])


wb = Workbook()
ws = wb.active

ws.cell(row=1, column=2, value = "Целей достигнуто")
ws.cell(row=1, column=3, value = "Кол-во пользователей")

for i in range (len(colNames)):
    ws.cell(row=i+2, column=1, value = colNames[i])
    ws.cell(row=i+2, column=2, value = colGoalReached[i])
    ws.cell(row=i+2, column=3, value = colUsers[i])

wb.save("file.xlsx")




##parsing response
#print ("HELLO")
#r = urllib.request.urlopen(req).read()
#print ("HE@")

#cont = json.loads(r.decode('utf-8'))
#counter = 0


#print (cont[0]['address']['city'])


#print (cont['data'][0]['dimensions'][0]['name'])

#print (cont['data'][0]['metrics'][0])

#AQAAAAAfG7AEAAUtKjM8zFheiEHQpFBiD0qG6Kg


